import openpyxl
from models import get_db, init_db, RESERVED_SUFFIXES


def import_excel(excel_path, db_path):
    """从 Excel 导入数据到 SQLite"""
    init_db(db_path)
    conn = get_db(db_path)

    wb = openpyxl.load_workbook(excel_path, read_only=True)

    for sort_order, sheet_name in enumerate(wb.sheetnames):
        ws = wb[sheet_name]
        title_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
        area_name = str(title_row[0] or sheet_name)

        cidr = f"{sheet_name}/24"

        conn.execute(
            "INSERT OR REPLACE INTO subnets (cidr, name, sort_order) VALUES (?, ?, ?)",
            (cidr, area_name, sort_order)
        )
        subnet_id = conn.execute(
            "SELECT id FROM subnets WHERE cidr = ?", (cidr,)
        ).fetchone()['id']

        header_row = next(ws.iter_rows(min_row=2, max_row=2, values_only=True))
        headers = [str(h).strip() if h else '' for h in header_row]
        has_mac = 'MAC地址' in headers or 'MAC' in ''.join(headers).upper()

        col_map = _build_column_map(headers, has_mac)

        for row in ws.iter_rows(min_row=3, max_row=ws.max_row, values_only=True):
            if not row or not row[1]:
                continue

            ip_address = str(row[1]).strip()
            if not ip_address:
                continue

            parts = ip_address.split('.')
            if len(parts) != 4:
                continue
            ip_suffix = int(parts[3])

            department = _cell_val(row, col_map.get('department'))
            username = _cell_val(row, col_map.get('username'))
            device = _cell_val(row, col_map.get('device'))
            device_model = _cell_val(row, col_map.get('device_model'))
            mac_address = _cell_val(row, col_map.get('mac_address'))
            location = _cell_val(row, col_map.get('location'))
            remark = _cell_val(row, col_map.get('remark'))

            if ip_suffix in RESERVED_SUFFIXES:
                status = 'reserved'
            elif department or username or device:
                status = 'occupied'
            else:
                status = 'free'

            conn.execute("""
                INSERT OR REPLACE INTO ip_addresses
                (subnet_id, ip_address, ip_suffix, status, department, username,
                 device, device_model, mac_address, location, remark)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """, (subnet_id, ip_address, ip_suffix, status, department, username,
                  device, device_model, mac_address, location, remark))

    conn.commit()
    conn.close()
    wb.close()

    total = _count_imported(db_path)
    return total


def _build_column_map(headers, has_mac):
    """根据表头动态映射列索引"""
    col_map = {}
    col_map['department'] = 2
    col_map['username'] = 3
    col_map['device'] = 4
    col_map['device_model'] = 5

    if has_mac:
        col_map['mac_address'] = 6
        col_map['location'] = 7
        col_map['remark'] = 8
    else:
        col_map['mac_address'] = None
        col_map['location'] = 6
        col_map['remark'] = 7

    return col_map


def _cell_val(row, idx):
    if idx is None or idx >= len(row):
        return ''
    val = row[idx]
    return str(val).strip() if val else ''


def _count_imported(db_path):
    conn = get_db(db_path)
    row = conn.execute("SELECT count(*) as cnt FROM ip_addresses").fetchone()
    conn.close()
    return row['cnt']
