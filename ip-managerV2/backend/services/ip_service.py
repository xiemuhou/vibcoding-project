from datetime import datetime, timezone
from models import get_db, RESERVED_SUFFIXES


def get_all_subnets(db_path):
    conn = get_db(db_path)
    subnets = conn.execute("""
        SELECT s.*,
            (SELECT count(*) FROM ip_addresses WHERE subnet_id = s.id) as total,
            (SELECT count(*) FROM ip_addresses WHERE subnet_id = s.id AND status = 'free') as free_count,
            (SELECT count(*) FROM ip_addresses WHERE subnet_id = s.id AND status = 'occupied') as occupied_count,
            (SELECT count(*) FROM ip_addresses WHERE subnet_id = s.id AND status = 'reserved') as reserved_count
        FROM subnets s ORDER BY s.sort_order
    """).fetchall()
    conn.close()
    return [dict(s) for s in subnets]


def get_subnet_ips(db_path, subnet_id):
    conn = get_db(db_path)
    ips = conn.execute("""
        SELECT * FROM ip_addresses
        WHERE subnet_id = ? ORDER BY ip_suffix
    """, (subnet_id,)).fetchall()
    conn.close()
    return [dict(ip) for ip in ips]


def get_ip_detail(db_path, ip_id):
    conn = get_db(db_path)
    ip = conn.execute("SELECT * FROM ip_addresses WHERE id = ?", (ip_id,)).fetchone()
    conn.close()
    return dict(ip) if ip else None


def occupy_ip(db_path, ip_id, data):
    conn = get_db(db_path)
    ip = conn.execute("SELECT * FROM ip_addresses WHERE id = ?", (ip_id,)).fetchone()
    if not ip:
        conn.close()
        return False, "IP不存在"
    if ip['status'] == 'reserved':
        conn.close()
        return False, "该IP为保留地址，不可占用"
    if ip['status'] == 'occupied':
        conn.close()
        return False, "该IP已被占用"

    conn.execute("""
        UPDATE ip_addresses SET
            status = 'occupied',
            department = ?, username = ?, device = ?,
            device_model = ?, mac_address = ?, location = ?,
            remark = ?, updated_at = ?
        WHERE id = ?
    """, (
        data.get('department', ''),
        data.get('username', ''),
        data.get('device', ''),
        data.get('device_model', ''),
        data.get('mac_address', ''),
        data.get('location', ''),
        data.get('remark', ''),
        datetime.now(timezone.utc).isoformat(),
        ip_id
    ))
    conn.commit()
    conn.close()
    return True, "占用成功"


def release_ip(db_path, ip_id):
    conn = get_db(db_path)
    ip = conn.execute("SELECT * FROM ip_addresses WHERE id = ?", (ip_id,)).fetchone()
    if not ip:
        conn.close()
        return False, "IP不存在"
    if ip['status'] == 'reserved':
        conn.close()
        return False, "该IP为保留地址，不可释放"
    if ip['status'] == 'free':
        conn.close()
        return False, "该IP已是空闲状态"

    conn.execute("""
        UPDATE ip_addresses SET
            status = 'free',
            department = '', username = '', device = '',
            device_model = '', mac_address = '', location = '',
            remark = '', updated_at = ?
        WHERE id = ?
    """, (datetime.now(timezone.utc).isoformat(), ip_id))
    conn.commit()
    conn.close()
    return True, "释放成功"


def update_ip(db_path, ip_id, data):
    conn = get_db(db_path)
    ip = conn.execute("SELECT * FROM ip_addresses WHERE id = ?", (ip_id,)).fetchone()
    if not ip:
        conn.close()
        return False, "IP不存在"
    if ip['status'] == 'reserved':
        conn.close()
        return False, "该IP为保留地址，不可编辑"

    conn.execute("""
        UPDATE ip_addresses SET
            department = ?, username = ?, device = ?,
            device_model = ?, mac_address = ?, location = ?,
            remark = ?, updated_at = ?
        WHERE id = ?
    """, (
        data.get('department', ip['department']),
        data.get('username', ip['username']),
        data.get('device', ip['device']),
        data.get('device_model', ip['device_model']),
        data.get('mac_address', ip['mac_address']),
        data.get('location', ip['location']),
        data.get('remark', ip['remark']),
        datetime.now(timezone.utc).isoformat(),
        ip_id
    ))
    conn.commit()
    conn.close()
    return True, "更新成功"


def search_ips(db_path, keyword):
    conn = get_db(db_path)
    like = f"%{keyword}%"
    results = conn.execute("""
        SELECT ip.*, s.cidr, s.name as subnet_name
        FROM ip_addresses ip
        JOIN subnets s ON ip.subnet_id = s.id
        WHERE ip.ip_address LIKE ?
           OR ip.username LIKE ?
           OR ip.department LIKE ?
           OR ip.mac_address LIKE ?
           OR ip.device LIKE ?
           OR ip.location LIKE ?
        ORDER BY ip.ip_address
        LIMIT 50
    """, (like, like, like, like, like, like)).fetchall()
    conn.close()
    return [dict(r) for r in results]


def export_excel(db_path):
    """导出数据为 openpyxl Workbook 对象"""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    conn = get_db(db_path)
    subnets = conn.execute("SELECT * FROM subnets ORDER BY sort_order").fetchall()

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    header_font = Font(bold=True)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font_white = Font(bold=True, color="FFFFFF")
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    for subnet in subnets:
        cidr = subnet['cidr']
        sheet_name = cidr.replace('/24', '')
        ws = wb.create_sheet(title=sheet_name)

        ws.merge_cells('A1:I1')
        ws['A1'] = subnet['name']
        ws['A1'].font = Font(bold=True, size=14)

        headers = ['序号', 'IP地址', '使用部门', '使用人', '使用设备', '设备型号', 'MAC地址', '位置', '备注']
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=2, column=col, value=h)
            cell.font = header_font_white
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
            cell.border = thin_border

        ips = conn.execute("""
            SELECT * FROM ip_addresses WHERE subnet_id = ? ORDER BY ip_suffix
        """, (subnet['id'],)).fetchall()

        for row_idx, ip in enumerate(ips, 3):
            values = [
                ip['ip_suffix'], ip['ip_address'], ip['department'],
                ip['username'], ip['device'], ip['device_model'],
                ip['mac_address'], ip['location'], ip['remark']
            ]
            for col, val in enumerate(values, 1):
                cell = ws.cell(row=row_idx, column=col, value=val)
                cell.border = thin_border

        ws.column_dimensions['A'].width = 6
        ws.column_dimensions['B'].width = 16
        ws.column_dimensions['C'].width = 14
        ws.column_dimensions['D'].width = 10
        ws.column_dimensions['E'].width = 18
        ws.column_dimensions['F'].width = 20
        ws.column_dimensions['G'].width = 20
        ws.column_dimensions['H'].width = 14
        ws.column_dimensions['I'].width = 20

    conn.close()
    return wb
