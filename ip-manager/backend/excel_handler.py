"""Excel 读写核心模块 — 处理 IP地址信息汇总.xlsx"""

import math
import os
import threading
import time
from pathlib import Path
from typing import Optional

import openpyxl
from openpyxl.utils import get_column_letter

from config import (
    EXCEL_PATH,
    READONLY_SHEETS,
    SKIP_SHEETS,
    DUAL_TABLE_SHEET,
    DUAL_TABLE_LEFT_COLS,
    WRITE_RETRY_COUNT,
    WRITE_RETRY_DELAY,
)

# ── 全局状态 ────────────────────────────────────────────

_lock = threading.Lock()
_writing = False  # 后台写入进行中，check_reload 需跳过
_pending_syncs = 0  # 尚未同步到 Excel 的缓存变更数
_last_mtime: float = 0.0
_cache: dict[str, list[dict]] = {}  # sheet_name -> list of ip records
_sheet_headers: dict[str, list[str]] = {}  # sheet_name -> column headers (normalized)
_sheet_row_offset: dict[str, int] = {}  # sheet_name -> data start row (1-indexed)


# ── 列映射工具 ──────────────────────────────────────────

# 每个 sheet 中标准列名 → 实际列名（用于统一查询）
_COLUMN_ALIASES = {
    "ip": ["ip地址", "ip address", "ip地址/端口号"],
    "department": ["使用部门"],
    "user": ["使用人"],
    "device": ["使用设备"],
    "model": ["设备型号"],
    "mac": ["mac地址", "mac"],
    "location": ["位置"],
    "remark": ["备注"],
    "serial": ["序号", "no."],
    "status": ["状态", "ip地址确认"],
}


def _normalize(name: str) -> str:
    """列名归一化：去空格、去换行、小写。"""
    return name.strip().replace("\n", "").replace("\r", "").lower()


def _find_column(headers: list[str], alias_group: list[str]) -> Optional[int]:
    """在表头中找到匹配的列索引。"""
    normalized_headers = [_normalize(h) for h in headers]
    for alias in alias_group:
        for i, h in enumerate(normalized_headers):
            if alias in h:
                return i
    return None


def _sheet_columns(headers: list[str]) -> dict[str, Optional[int]]:
    """返回 sheet 的标准列索引映射。"""
    return {
        "serial": _find_column(headers, _COLUMN_ALIASES["serial"]),
        "ip": _find_column(headers, _COLUMN_ALIASES["ip"]),
        "department": _find_column(headers, _COLUMN_ALIASES["department"]),
        "user": _find_column(headers, _COLUMN_ALIASES["user"]),
        "device": _find_column(headers, _COLUMN_ALIASES["device"]),
        "model": _find_column(headers, _COLUMN_ALIASES["model"]),
        "mac": _find_column(headers, _COLUMN_ALIASES["mac"]),
        "location": _find_column(headers, _COLUMN_ALIASES["location"]),
        "remark": _find_column(headers, _COLUMN_ALIASES["remark"]),
        "status": _find_column(headers, _COLUMN_ALIASES["status"]),
    }


def _is_free(row_data: dict) -> bool:
    """判断 IP 是否空闲：使用人为空。"""
    user = row_data.get("useUser", "")
    if user is None:
        return True
    return str(user).strip() == ""


# ── 读取 ────────────────────────────────────────────────


def load_all_sheets(filepath: Optional[str] = None) -> dict[str, list[dict]]:
    """加载所有 Sheet 到缓存。返回 {sheet_name: [records]}。"""
    global _last_mtime, _cache, _sheet_headers, _sheet_row_offset

    path = Path(filepath or EXCEL_PATH)
    if not path.exists():
        raise FileNotFoundError(f"Excel 文件不存在: {path}")

    _last_mtime = path.stat().st_mtime
    _cache = {}
    _sheet_headers = {}
    _sheet_row_offset = {}

    wb = openpyxl.load_workbook(path, data_only=True)

    for sheet_name in wb.sheetnames:
        if sheet_name in SKIP_SHEETS:
            continue

        ws = wb[sheet_name]
        records, headers, data_start = _parse_sheet(sheet_name, ws)
        if records:
            _cache[sheet_name] = records
            _sheet_headers[sheet_name] = headers
            _sheet_row_offset[sheet_name] = data_start

    wb.close()
    return _cache


def _parse_sheet(sheet_name: str, ws) -> tuple[list[dict], list[str], int]:
    """解析单个 Sheet，返回 (记录列表, 表头列表, 数据起始行号)。"""
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return [], [], 1

    # 检测表头行：大多数 Sheet 第 1 行是标题，第 2 行是列名
    header_row_idx = 0
    data_start = 1

    # 查看前两行，找到真正的列名行
    for check_idx in range(min(3, len(rows))):
        row = rows[check_idx]
        if row is None:
            continue
        # 检查是否包含"IP地址"或"序号"关键词
        text = " ".join(str(c) for c in row if c is not None).lower()
        if "ip地址" in text or "ip address" in text:
            header_row_idx = check_idx
            data_start = check_idx + 1
            break

    # 如果没找到 IP 地址列，尝试第 2 行（0-indexed: 1）
    if data_start == 1 and len(rows) > 1:
        header_row_idx = min(1, len(rows) - 1)
        data_start = header_row_idx + 1

    # 获取表头
    header_row = rows[header_row_idx]
    raw_headers = [str(c).strip() if c is not None else "" for c in header_row]

    # 10.10.128.0 双表处理：只取左侧表
    if sheet_name == DUAL_TABLE_SHEET:
        raw_headers = raw_headers[:DUAL_TABLE_LEFT_COLS]

    col_map = _sheet_columns(raw_headers)
    ip_col = col_map["ip"]

    # 如果没有 IP 地址列，返回空
    if ip_col is None:
        return [], raw_headers, data_start + 1

    records = []
    for row_idx in range(data_start, len(rows)):
        row = rows[row_idx]
        if row is None:
            continue

        # 转为 list 便于索引
        values = list(row) if row else []
        if sheet_name == DUAL_TABLE_SHEET:
            values = values[:DUAL_TABLE_LEFT_COLS]

        # 获取 IP 地址
        if ip_col >= len(values):
            continue
        ip_val = values[ip_col]
        if ip_val is None:
            # 有些 Sheet 中空行的 IP 是预设的，检查后续列
            continue
        ip_str = str(ip_val).strip()
        if not ip_str or ip_str == "None":
            continue

        record = {
            "excel_row": row_idx + 1,  # Excel 行号（1-indexed）
            "sheet": sheet_name,
            "ip": ip_str,
            "department": _cell_str(values, col_map["department"]),
            "useUser": _cell_str(values, col_map["user"]),
            "useDevice": _cell_str(values, col_map["device"]),
            "model": _cell_str(values, col_map["model"]),
            "macAddress": _cell_str(values, col_map["mac"]),
            "location": _cell_str(values, col_map["location"]),
            "remark": _cell_str(values, col_map["remark"]),
            "serial": _cell_str(values, col_map["serial"]),
            "status": _cell_str(values, col_map["status"]),
            "free": True,  # 下面重算
        }
        record["free"] = _is_free(record)
        records.append(record)

    return records, raw_headers, data_start + 1  # Excel 行号从 1 开始


def _cell_str(values: list, idx: Optional[int]) -> str:
    """安全获取单元格字符串值。"""
    if idx is None or idx >= len(values):
        return ""
    v = values[idx]
    if v is None:
        return ""
    return str(v).strip()


# ── 缓存访问 ────────────────────────────────────────────


def check_reload() -> bool:
    """检测 Excel 是否被外部修改，自动重载。后台写入期间或有待同步变更时跳过。"""
    if _writing or _pending_syncs > 0:
        return False
    path = Path(EXCEL_PATH)
    if not path.exists():
        return False
    try:
        current_mtime = path.stat().st_mtime
        if current_mtime != _last_mtime:
            load_all_sheets()
            return True
    except Exception as e:
        print(f"[WARN] check_reload 失败: {e}")
    return False


def get_cached_sheets() -> dict[str, list[dict]]:
    """获取缓存的 Sheet 数据。"""
    check_reload()
    return _cache


def get_sheet_list() -> list[dict]:
    """获取 Sheet 概览列表。"""
    check_reload()
    result = []
    for name, records in _cache.items():
        total = len(records)
        used = sum(1 for r in records if not r["free"])
        free = total - used
        result.append({
            "name": name,
            "total": total,
            "used": used,
            "free": free,
            "usageRate": round(used / total * 100, 1) if total > 0 else 0,
            "readonly": name in READONLY_SHEETS,
        })
    return result


def get_sheet_records(sheet_name: str) -> list[dict]:
    """获取指定 Sheet 的所有 IP 记录。"""
    check_reload()
    return _cache.get(sheet_name, [])


# ── 写入（缓存层） ──────────────────────────────────────


def occupy_ip_cache(
    sheet_name: str,
    excel_row: int,
    use_user: str,
    department: str = "",
    use_device: str = "",
    model: str = "",
    mac_address: str = "",
    location: str = "",
    remark: str = "",
) -> bool:
    """占用 IP：立即更新内存缓存，返回 True 成功 / False 冲突。"""
    with _lock:
        records = _cache.get(sheet_name, [])
        target = None
        for r in records:
            if r["excel_row"] == excel_row:
                target = r
                break
        if target is None:
            raise ValueError(f"未找到行: sheet={sheet_name}, row={excel_row}")
        if not target["free"]:
            return False

        target["department"] = department
        target["useUser"] = use_user
        target["useDevice"] = use_device
        target["model"] = model
        target["macAddress"] = mac_address
        target["location"] = location
        target["remark"] = remark
        target["free"] = False
        return True


def release_ip_cache(sheet_name: str, excel_row: int) -> bool:
    """释放 IP：立即更新内存缓存，返回 True 成功 / False 已空闲。"""
    with _lock:
        records = _cache.get(sheet_name, [])
        target = None
        for r in records:
            if r["excel_row"] == excel_row:
                target = r
                break
        if target is None:
            raise ValueError(f"未找到行: sheet={sheet_name}, row={excel_row}")
        if target["free"]:
            return False

        target["department"] = ""
        target["useUser"] = ""
        target["useDevice"] = ""
        target["model"] = ""
        target["macAddress"] = ""
        target["location"] = ""
        target["remark"] = ""
        target["status"] = ""
        target["free"] = True
        return True


# ── 写入（Excel 层）─ 后台线程执行 ──────────────────────

import queue

_write_queue = queue.Queue()
_writer_started = False


def _write_cell(ws, row: int, col: int, value) -> None:
    ws.cell(row=row, column=col).value = value if value else ""


def _safe_write(ws, row: int, col: Optional[int], value) -> None:
    if col is None or col < 0:
        return
    _write_cell(ws, row, col + 1, value)


def _retry_open(path: Path) -> openpyxl.Workbook:
    """带重试打开 Excel 用于写入（保留公式和样式）。"""
    last_error = None
    for i in range(WRITE_RETRY_COUNT):
        try:
            return openpyxl.load_workbook(path)
        except PermissionError as e:
            last_error = e
            if i < WRITE_RETRY_COUNT - 1:
                time.sleep(WRITE_RETRY_DELAY)
    raise PermissionError(f"Excel 文件被占用，请关闭后重试 ({WRITE_RETRY_COUNT} 次重试均失败)") from last_error


def schedule_sync(sheet_name: str, excel_row: int) -> None:
    """将 Excel 写入任务加入后台队列，立即返回（微秒级）。"""
    global _writer_started, _pending_syncs
    if not _writer_started:
        _start_writer()
    _pending_syncs += 1
    _write_queue.put((sheet_name, excel_row))


def _start_writer():
    """启动后台写入线程。"""
    global _writer_started
    _writer_started = True
    t = threading.Thread(target=_writer_loop, daemon=True, name="excel-writer")
    t.start()


def _writer_loop():
    """后台线程：从队列取任务，写入 Excel。"""
    while True:
        try:
            sheet_name, excel_row = _write_queue.get()
        except Exception:
            continue
        if sheet_name is None:
            break
        try:
            _do_sync(sheet_name, excel_row)
        except Exception as e:
            print(f"[ERROR] Excel 写入失败 ({sheet_name} row={excel_row}): {e}")


def _do_sync(sheet_name: str, excel_row: int):
    """执行单次 Excel 写入（在后台线程中调用）。

    锁只保护缓存读取，Excel I/O 在锁外执行。
    _pending_syncs 在顶层 finally 中递减，确保所有路径只减一次。
    """
    global _writing, _last_mtime, _pending_syncs

    try:
        path = Path(EXCEL_PATH)
        if not path.exists():
            return

        # 锁内：仅复制需要的数据
        with _lock:
            records = _cache.get(sheet_name, [])
            target = None
            for r in records:
                if r["excel_row"] == excel_row:
                    target = dict(r)
                    break
            if target is None:
                return
            headers = list(_sheet_headers.get(sheet_name, []))

        # 锁外：慢速 Excel I/O
        col_map = _sheet_columns(headers)
        wb = _retry_open(path)
        try:
            ws = wb[sheet_name]

            if target["free"]:
                for key in ["department", "user", "device", "model", "mac", "location", "remark", "status"]:
                    _safe_write(ws, excel_row, col_map.get(key), "")
            else:
                _safe_write(ws, excel_row, col_map.get("department"), target.get("department"))
                _safe_write(ws, excel_row, col_map.get("user"), target.get("useUser"))
                _safe_write(ws, excel_row, col_map.get("device"), target.get("useDevice"))
                _safe_write(ws, excel_row, col_map.get("model"), target.get("model"))
                _safe_write(ws, excel_row, col_map.get("mac"), target.get("macAddress"))
                _safe_write(ws, excel_row, col_map.get("location"), target.get("location"))
                _safe_write(ws, excel_row, col_map.get("remark"), target.get("remark"))

            _writing = True
            try:
                wb.save(str(path))
                with _lock:
                    _last_mtime = path.stat().st_mtime
            finally:
                _writing = False
        finally:
            wb.close()
    finally:
        _pending_syncs = max(0, _pending_syncs - 1)


# ── 启动初始化 ──────────────────────────────────────────


def init_excel_handler():
    """启动时加载 Excel。"""
    path = Path(EXCEL_PATH)
    if not path.exists():
        print(f"[WARN] Excel 文件不存在: {path}，将以空数据启动")
        return
    load_all_sheets()
    print(f"[OK] Excel 加载完成: {len(_cache)} 个 Sheet, {sum(len(v) for v in _cache.values())} 条 IP 记录")
